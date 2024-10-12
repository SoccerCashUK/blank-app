import streamlit as st
import pandas as pd
import statsmodels.api as sm
import statsmodels.formula.api as smf
import numpy as np
from scipy.stats import poisson
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

st.title('Germany 2 League 2024/2025 Prediction App')
st.write('Predict the outcome of Germany 2 League matches using statistical models.')


#st.sidebar.markdown('Catch up with the latest Chelsea FC news at <a href="https://www.footballcatchup.com" target="_self">Football Catchup</a>', unsafe_allow_html=True)



home = st.sidebar.selectbox('Select home team', ['Braunschweig', 'Darmstadt', 'Elversberg', 'FC Koln', 
                                                 'Fortuna Dusseldorf', 'Greuther Furth', 'Hamburg', 'Hannover', 'Hertha', 
                                                 'Kaiserslautern', 'Karlsruhe', 'Magdeburg', 'Nurnberg', 'Paderborn', 
                                                 'Preußen Münster', 'Regensburg', 'Schalke 04', 'Ulm'])

away = st.sidebar.selectbox('Select away team', ['Braunschweig', 'Darmstadt', 'Elversberg', 'FC Koln', 
                                                 'Fortuna Dusseldorf', 'Greuther Furth', 'Hamburg', 'Hannover', 'Hertha', 
                                                 'Kaiserslautern', 'Karlsruhe', 'Magdeburg', 'Nurnberg', 'Paderborn', 
                                                 'Preußen Münster', 'Regensburg', 'Schalke 04', 'Ulm'])

button = st.sidebar.button('Predict')

@st.cache_data
def load_data():
    try:
        data = pd.read_csv('https://www.football-data.co.uk/mmz4281/2425/D2.csv')
        epl = data[['HomeTeam', 'AwayTeam', 'FTHG', 'FTAG']]
        epl = epl.rename(columns={'FTHG': 'HomeGoals', 'FTAG': 'AwayGoals'})
        home_data = epl.iloc[:, 0:3].assign(home=1).rename(columns={'HomeTeam': 'team', 'AwayTeam': 'opponent', 'HomeGoals': 'goals'})
        away_data = epl.iloc[:, [1, 0, 3]].assign(home=0).rename(columns={'AwayTeam': 'team', 'HomeTeam': 'opponent', 'AwayGoals': 'goals'})
        df = pd.concat([home_data, away_data])
        return df
    except Exception as e:
        st.error(f"Error loading data: {e}")
        return pd.DataFrame()

df = load_data()

@st.cache_resource
def build_model(data):
    formula = 'goals ~ team + opponent + home'
    model = smf.glm(formula=formula, data=data, family=sm.families.Poisson()).fit()
    return model

# Monte Carlo Simulation function
def monte_carlo_simulation(home_team, away_team, model, n_simulations=10000):
    home_goals_pred = model.predict(pd.DataFrame({'team': [home_team], 'opponent': [away_team], 'home': [1]})).values[0]
    away_goals_pred = model.predict(pd.DataFrame({'team': [away_team], 'opponent': [home_team], 'home': [0]})).values[0]
    
    home_wins, draws, away_wins = 0, 0, 0
    
    for _ in range(n_simulations):
        home_goals_sim = np.random.poisson(home_goals_pred)
        away_goals_sim = np.random.poisson(away_goals_pred)
        
        if home_goals_sim > away_goals_sim:
            home_wins += 1
        elif home_goals_sim < away_goals_sim:
            away_wins += 1
        else:
            draws += 1
    
    total_simulations = home_wins + draws + away_wins
    home_win_prob = home_wins / total_simulations * 100
    draw_prob = draws / total_simulations * 100
    away_win_prob = away_wins / total_simulations * 100
    
    return round(home_win_prob), round(draw_prob), round(away_win_prob)

# Function to save predictions to Excel
def save_to_excel(home_team, away_team, home_goals, away_goals, home_win_prob, draw_prob, away_win_prob, file_name="predictions.xlsx"):
    today_date = datetime.today().strftime('%Y-%m-%d')  # Get current date
    
    # Create new data with League column
    new_data = pd.DataFrame({
        'Date': [today_date],
        'League': ['Germany 2'],  # Add League column here
        'Home Team': [home_team],
        'Away Team': [away_team],
        'Home Goals Predicted': [home_goals],
        'Away Goals Predicted': [away_goals],
        'Home Win Probability (%)': [home_win_prob],
        'Draw Probability (%)': [draw_prob],
        'Away Win Probability (%)': [away_win_prob]
    })
    
    if os.path.exists(file_name):
        # If the file exists, append to the existing file
        with pd.ExcelWriter(file_name, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
            new_data.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        # Center-align all data in the Excel sheet
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        wb.save(file_name)
    else:
        # If the file does not exist, create a new file and center-align all data
        with pd.ExcelWriter(file_name, mode='w', engine='openpyxl') as writer:
            new_data.to_excel(writer, index=False)
        wb = load_workbook(file_name)
        ws = wb.active
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
        wb.save(file_name)

# Building and predicting based on the model
if not df.empty:
    model = build_model(df)

    home_goals = int(model.predict(pd.DataFrame({'team': [home], 'opponent': [away], 'home': [1]})).iloc[0])
    away_goals = int(model.predict(pd.DataFrame({'team': [away], 'opponent': [home], 'home': [0]})).iloc[0])
    
    if button:
        if home == away:
            st.error("You can't predict the same team.")
        else:
            st.write(f"Score Prediction: {home} {home_goals} - {away_goals} {away}")
            
            home_win_prob, draw_prob, away_win_prob = monte_carlo_simulation(home, away, model, n_simulations=10000)
            
            st.write(f"Monte Carlo Simulation Results (10,000 simulations):")
            st.write(f"Prediction for Home win: {home_win_prob}%")
            st.write(f"Prediction for Draw: {draw_prob}%")
            st.write(f"Prediction for Away win: {away_win_prob}%")
            
            # Save predictions to Excel
            save_to_excel(home, away, home_goals, away_goals, home_win_prob, draw_prob, away_win_prob)
else:
    st.error("No data available to build the model.")
